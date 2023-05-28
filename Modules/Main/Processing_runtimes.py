import openpyxl
import time
import numpy as np


def data_analyse(slopes_mask_in_black_and_white: np.ndarray, km_radius: float, st: float, cputime_start: float) -> None:
    # count number of 255 in slopes_mask_in_black_and_white
    count_slopes_good = np.count_nonzero(slopes_mask_in_black_and_white == 255)
    slopy = round(100 * count_slopes_good / slopes_mask_in_black_and_white.size, 2)
    area = (2 * km_radius) ** 2
    total_time = time.time() - st
    cpu_total_time = time.process_time() - cputime_start
    # write to excel
    print(slopy, area, total_time)
    save_result_to_excel(slopy, area, total_time, cpu_total_time)


def save_result_to_excel(slopy: float, area: float, total_time: float, cpu_total_time: float) -> None:
    workbook = openpyxl.load_workbook("results.xlsx")
    worksheet = workbook.active
    worksheet.title = "result"
    worksheet.cell(row=1, column=1, value="slopy%")
    worksheet.cell(row=1, column=2, value="area [km^2]")
    worksheet.cell(row=1, column=3, value="total time [sec]")
    worksheet.cell(row=1, column=4, value="cpu total time [sec]")
    # add to filename
    last_row = worksheet.max_row
    worksheet.cell(row=last_row + 1, column=1, value=slopy)
    worksheet.cell(row=last_row + 1, column=2, value=area)
    worksheet.cell(row=last_row + 1, column=3, value=total_time)
    worksheet.cell(row=last_row + 1, column=4, value=cpu_total_time)

    workbook.save("results.xlsx")
