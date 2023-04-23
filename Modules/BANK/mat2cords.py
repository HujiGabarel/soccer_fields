import openpyxl


def save_to_excel(filename, helipadList):
    # Load the existing workbook or create a new one if the file doesn't exist
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the worksheet in the workbook or create a new one if it doesn't exist
    worksheet_name = "Sheet"
    if worksheet_name not in workbook.sheetnames:
        workbook.create_sheet(worksheet_name)
    worksheet = workbook[worksheet_name]

    # Write the column names to the first row of the worksheet
    column_names = ["X Coordinate", "Y Coordinate", "Size"]
    for i, column_name in enumerate(column_names):
        worksheet.cell(row=1, column=i+1, value=column_name)

    # Get the existing helipad coordinates from the worksheet
    existing_helipad_coords = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        existing_helipad_coords.add((row[0], row[1]))

    # Write the new helipad data to the worksheet only if it doesn't exist already
    for i, helipad in enumerate(helipadList):
        coords = helipad[0]
        size = helipad[1]
        if coords not in existing_helipad_coords:
            last_row = worksheet.max_row
            worksheet.cell(row=last_row+1, column=1, value=coords[0])
            worksheet.cell(row=last_row+1, column=2, value=coords[1])
            worksheet.cell(row=last_row+1, column=3, value=size)
            existing_helipad_coords.add(coords)

    # Save the workbook
    workbook.save(filename)




def mat2cords(satellitePicture, heliPadList, topLeftCords):
    helipadList = []  # The list of center of the helipads CORDS with size
    for i in heliPadList:
        center = MidOfHeliPad(i[0], i[1], i[2], i[3])  # insert the top left x and y cords then the bot_right cords.
        width = i[2] - i[0] + 1  # calculate the width of the submatrix
        height = i[3] - i[1] + 1  # calculate the height of the submatrix
        size = width * height  # calculate the size of the submatrix (area)

        # Calculate the real-world coordinates of the helipad center
        centerX, centerY = center
        realX = topLeftCords[0] + (centerX - 1)  # subtract 1 to convert from 1-based to 0-based indexing
        realY = topLeftCords[1] - (centerY - 1)  # subtract 1 and negate to convert from 1-based to 0-based indexing and flip y-axis
        realCenter = (realX, realY)

        helipadList.append((realCenter, size))

    return helipadList




def MidOfHeliPad(topLeftX, topLeftY, botRightX, botRightY):
    subMatX = botRightX - topLeftX  # length of helipad
    subMatY = botRightY - topLeftY  # width of helipad
    subMatMidX = subMatX / 2 + 1
    subMatMidY = subMatY / 2 + 1
    subMatMidX = int(subMatMidX)
    subMatMidY = int(subMatMidY)
    finalX = topLeftX + subMatMidX
    finalY = topLeftY + subMatMidY
    return finalX, finalY





def main():
    # Example inputs
    satellitePicture = [[1, 0, 0, 1, 1],
                        [1, 0, 0, 1, 1],
                        [1, 1, 1, 0, 1],
                        [0, 1, 1, 1, 1],
                        [0, 0, 1, 1, 1]]
    heliPadList = [(0, 2, 2, 2), (1, 1, 4, 4),(0, 2, 2, 2), (1, 1, 4, 4),(0, 2, 2, 4), (0, 0, 4, 4)]
    topLeftCords = (10, 10)

    # Call the mat2cords function
    helipadList = mat2cords(satellitePicture, heliPadList, topLeftCords)

    # Print the output of the mat2cords function
    print("Helipad List:")
    for helipad in helipadList:
        print(helipad)

    # Call the save_to_excel function
    filename = "helipads.xlsx"
    try:
        save_to_excel(filename, helipadList)
        print(f"Helipads saved to {filename}")
    except Exception as e:
        print(f"Error saving helipads to {filename}: {e}")

if __name__ == '__main__':
    main()